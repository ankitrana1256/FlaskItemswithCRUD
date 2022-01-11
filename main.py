import os
from flask import Flask, request, render_template, redirect
from flask.helpers import url_for
import openpyxl
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///products.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)


class products(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product = db.Column(db.String(20), nullable=False)
    prod_image = db.Column(db.String(400), nullable=False)
    desc = db.Column(db.String(400), nullable=False)
    date = db.Column(db.DateTime, default=datetime.now)


@app.route("/", methods=["POST", "GET"])
def homepage():
    if (request.method) == "POST":
        if request.form["submit"] == "submit":
            product = request.form["product"]
            prod_image = request.form["prod_image"]
            desc = request.form["desc"]
            new_product = products(
                product=product,
                prod_image=prod_image,
                desc=desc,
            )
            db.session.add(new_product)
            db.session.commit()
            return redirect(url_for("homepage"))

    ProdList = products.query.all()
    return render_template("homepage.html", prods=ProdList)


@app.route("/process", methods=["POST", "GET"])
def process():
    if request.method == "POST":
        print("Here")
        file = request.files["file"]
        path = f"static/{file.filename}"
        file.save(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        max_data = ws.max_row
        for i in range(2, max_data + 1):
            itemname = ws.cell(column=1, row=i).value
            description = ws.cell(column=2, row=i).value
            url = ws.cell(column=3, row=i).value
            new_prod = products(product=itemname, desc=description, prod_image=url)
            db.session.add(new_prod)
            db.session.commit()
        return redirect(url_for("homepage"))


@app.route("/delete/<int:id>")
def delete(id):
    product = products.query.filter_by(id=id).first()
    db.session.delete(product)
    db.session.commit()
    return redirect(url_for("homepage"))


@app.route("/update/<int:id>", methods=["POST", "GET"])
def update(id):
    prod = products.query.filter_by(id=id).first()
    if (request.method) == "POST":
        product = request.form["product"]
        prod_image = request.form["prod_image"]
        desc = request.form["desc"]
        prod.product = product
        prod.prod_image = prod_image
        prod.desc = desc
        db.session.add(prod)
        db.session.commit()
        return redirect(url_for("homepage"))
    return render_template("update.html", prod=prod)


if __name__ == "__main__":
    app.run(debug=True, port="8000")
